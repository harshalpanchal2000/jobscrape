"""Excel export for scraped job data."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def export_to_excel(jobs):
    """Create a styled Excel workbook from job data and return as bytes.

    Args:
        jobs: List of job dicts with keys: title, company, location, posted, url, description.

    Returns:
        BytesIO buffer containing the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn Jobs"

    # Header styling (LinkedIn blue)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0A66C2", end_color="0A66C2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["#", "Job Title", "Company", "Location", "Posted Date", "Job URL", "Job Description"]
    col_widths = [5, 35, 25, 25, 15, 55, 80]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    wrap_align = Alignment(wrap_text=True, vertical="top")
    link_font = Font(color="0563C1", underline="single")

    for row_idx, job in enumerate(jobs, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=job.get("title", ""))
        ws.cell(row=row_idx, column=3, value=job.get("company", ""))
        ws.cell(row=row_idx, column=4, value=job.get("location", ""))
        ws.cell(row=row_idx, column=5, value=job.get("posted", ""))

        url_cell = ws.cell(row=row_idx, column=6, value=job.get("url", ""))
        url_cell.font = link_font
        if job.get("url"):
            url_cell.hyperlink = job["url"]

        desc_cell = ws.cell(row=row_idx, column=7, value=job.get("description", ""))
        desc_cell.alignment = wrap_align

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
